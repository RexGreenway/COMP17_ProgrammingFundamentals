# Thomas Rex Greenway, 201198319

## IMPORTS ##
import openpyxl


### CARD CLASS ###
class Card():
    """
    Card is an individual trading card. Has Name, Type, HP, Moves (dict.), and Shiny (boolean). Also contains Card Average Damage as a private data tracker.
    """
    def __init__(self, theName, theType, theHP, theMoves, isShiny):
        self.theName = theName
        self.theType = theType
        self.theHP = theHP
        self.isShiny = isShiny
        self.theMoves = theMoves

        # Private variable
        self._avgDamage = self.getAverageDamage()
    
    def __str__(self):
        return "Card: {self.theName}\nType: {self.theType}\nHP: {self.theHP}\nMoves: {self.theMoves}\nShiny: {self.isShiny}".format(self=self)
    
    # Card Average Damage Getter
    def getAverageDamage(self):
        """
        Returns average damage for card across all of its moves.

        Parameters:
            Nothing
        Returns:
            float - mean average damage for card instance.
        """
        totalDamage =  0
        for move in self.theMoves:
            dmg = self.theMoves.get(move)
            totalDamage += dmg
        
        # Error handling for case where card has no moves. Cannot divide by zero.
        try:
            self._avgDamage = totalDamage/len(self.theMoves)
        except ZeroDivisionError:
            self._avgDamage = 0
        return self._avgDamage
        

### DECK CLASS ###
class Deck():
    """
    Deck is a collection of trading cards. Has Deck, a list to be populated with card instances, and 3 private data trackers, Deck Average Damage, Total Cards, Total Shiny Cards.
    """
    def __init__(self):
        self.deck = []

        # Private Variables
        self._avgDamage = 0
        self._totalCardNum = 0
        self._totalShinyNum = 0


    def inputFromFile(self, fileName):
        """
        Reads an .xlxs file and creates respective instances of Card class and add each of them, using addCard, to the deck list...

        Parameters:
            fileName: string - a file of <fileName>.xlxs type.
        Returns:
            Nothing
        """
        # Loads Workbook File 
        book = openpyxl.load_workbook(fileName + ".xlsx")
        sheet = book.active

        # Finds total number of entries (-1 because of openpyxl quirk)
        numRows = sheet.max_row - 1

        # Reads file and assigns values row by row, starting at row 2 to ignore column headings.
        for row in sheet.iter_rows(min_row = 2, max_row = numRows):
            theName = row[0].value
            theType = row[1].value
            theHP = row[2].value
            isShiny = bool(row[3].value)
            # Moves saved to a dictionary with key: moveName, value: damage.
            theMoves = {}
            for i in range(4, 13, 2):
                # Will break if card has <5 moves. 
                if (row[i].value == None):
                    break
                theMoves[row[i].value] = row[i + 1].value
            
            # Creation of card instances. Add them to the deck list.
            theName = Card(theName, theType, theHP, theMoves, isShiny)
            self.addCard(theName)


    def __str__(self):
        return "DECK:\n\tTotal Card Number: {self._totalCardNum}\n\tTotal Shiny Number: {self._totalShinyNum}\n\tAverage Damage Value: {self._avgDamage}".format(self=self)
    

    def addCard(self, theCard):
        """
        Adds a card instance to the deck list. Also updates relative data tracking variables; total card number, total shiny number, and average damage.

        Parameters:
            theCard: variable name of the desired card instance.
        Returns:
            Nothing
        """
        # Ensuring correct argument type.
        if isinstance(theCard, Card):
            self.deck.append(theCard)
            self._totalCardNum += 1
            if (theCard.isShiny == True):
                self._totalShinyNum += 1
            self._avgDamage = self.getAverageDamage()
        else:
            print("PLEASE INPUT A CARD")


    def rmCard(self, theCard):
        """
        Removes a card instance to the deck list. Also updates relative data tracking variables; total card num,ber, total shiny number, and avergae damage.

        Parameters:
            theCard: variable name of the desired card instance.
        Returns:
            Nothing
        """
        # Ensuring correct argument type.
        if isinstance(theCard, Card):
            self.deck.remove(theCard)
            self._totalCardNum -= 1
            if (theCard.isShiny == True):
                self._totalShinyNum -= 1
            self._avgDamage = self.getAverageDamage()
        else:
            print("PLEASE INPUT A CARD")


    def getMostPowerful(self):
        """
        Returns the most powerful card in the deck list. The greater the cards average damage the greater its power.

        Parameters:
            Nothing
        Returns:
            Card object - the most powerful card instance.
        """
        damage = 0
        for card in self.deck:
            if (card.getAverageDamage() > damage):
                damage = card.getAverageDamage()
                mostPowerful = card
        return mostPowerful


    # Deck Average Damage Getter
    def getAverageDamage(self):
        """
        Returns the average damage of all cards in the deck list.

        Parameters:
            Nothing
        Returns:
            float - the average damage of the all cards in the deck list.
        """
        totalDamage = 0
        for card in self.deck:
            dmg = card.getAverageDamage()
            totalDamage += dmg

        # Error handling for when the deck is empty. Cannot divide by zero.
        try:
            self._avgDamage = totalDamage/len(self.deck)
        except ZeroDivisionError:
            self._avgDamage = 0
        return round(self._avgDamage, 1) 


    def viewAllCards(self):
        """
        Prints information of all cards in the deck list.

        Parameters:
            Nothing
        Returns:
            string - card information from the deck list.
        """
        if (self._totalCardNum > 0):    
            print("THIS DECK CONTAINS " + str(self._totalCardNum) + " CARDS.")
            for card in self.deck:
                print("\n", str(card))
        else:
            print("THIS DECK IS EMPTY")


    def viewAllShinyCards(self):
        """
        Prints information of all shiny cards in the deck list.

        Parameters:
            Nothing
        Returns:
            string - card information from the deck list.
        """
        if (self._totalShinyNum > 0):
            print("THIS DECK CONTAINS " + str(self._totalShinyNum) + " SHINY CARDS:")
            for card in self.deck:
                if (card.isShiny == True):
                    print ("\n", str(card))
        else:
            print("THIS DECK CONTAINS NO SHINYS")


    def viewAllByType(self, theType):
        """
        Prints information of all cards with given type in the deck list.

        Parameters:
            theType: string -  from potential types; 'Magi', 'Water', 'Fire', 'Earth', 'Air',' Astral'.
        Returns:
            string - card information from the deck list.
        """
        if (theType in ["Magi", "Water", "Fire", "Earth", "Air", "Astral"]):
            print("THIS DECK CONTAINS CARD OF TYPE: " + theType)
            for card in self.deck:
                if (card.theType == theType):
                    print("\n", str(card))
        else:
            print("THERE IS NO TYPE OF THAT NAME")


    def getCards(self):
        """
        Returns the deck list.

        Parameters:
            Nothing 
        Returns:
            list - list of card instances from the deck list.
        """
        return self.deck
    

    def saveToFile(self, fileName):
        """
        Saves a new file of the deck to <fileName>.xlsx.

        Parameters:
            fileName: string - desired file name for deck file.
        Returns:
            Nothing
        """
        # Start a new workbook.
        book = openpyxl.Workbook()
        sheet = book.active

        # Establish column names as first row of deck file.
        firstRow = ("Name",
            "Type",
            "HP",
            "Shiny",
            "Move Name 1",
            "Damage 1",
            "Move Name 2",
            "Damage 2",
            "Move Name 3",
            "Damage 3",
            "Move Name 4",
            "Damage 4",
            "Move Name 5",
            "Damage 5")
        sheet.append(firstRow)

        # Create and append a tuple for each card in the deck.
        for card in self.deck:              
            cardTuple = [
                card.theName,
                card.theType,
                card.theHP,
                int(card.isShiny)]
            for move in card.theMoves.items():
                cardTuple.append(move[0])
                cardTuple.append(move[1])
            sheet.append(tuple(cardTuple))
        
        #Save workbook.
        book.save(fileName + ".xlsx")


    # PROPERTIES
    @property
    def totalCardNum(self):
        return self._totalCardNum

    @property
    def totalShinyNum(self):
        return self._totalShinyNum

    @property
    def avgDamage(self):
        return self._avgDamage
